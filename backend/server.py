"""B2B Marketplace API — Auth, Catalog, Cart, Orders, Vendors, RFQ, Admin."""
import os, uuid, logging, io
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, List

from fastapi import FastAPI, APIRouter, Depends, HTTPException, UploadFile, File, Request, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr, Field
from bson import ObjectId

from config import MONGO_URL, DB_NAME, DEFAULT_SETTINGS, UPLOAD_DIR, RAZORPAY_KEY_ID
from auth_utils import (
    hash_pw, verify_pw, make_access_token, make_refresh_token, decode_token,
    get_current_user, get_current_user_optional, require_permission, require_roles,
)
from integrations import (
    razor_create_order, razor_verify_signature, razor_verify_webhook,
    ship_create_order, ship_track_awb, send_email, price_for_qty,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
log = logging.getLogger("app")

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="B2B Marketplace API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_credentials=True, allow_methods=["*"], allow_headers=["*"],
)

api = APIRouter(prefix="/api")


def now():
    return datetime.now(timezone.utc).isoformat()


def new_id():
    return str(uuid.uuid4())


def clean(doc: dict) -> dict:
    if not doc:
        return doc
    doc.pop("_id", None)
    return doc


def clean_list(docs):
    return [clean(d) for d in docs]


# ============ MODELS ============
class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str
    phone: Optional[str] = ""
    company: Optional[str] = ""
    gstin: Optional[str] = ""
    role: str = "customer"  # customer | vendor


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class CategoryIn(BaseModel):
    name: str
    slug: Optional[str] = None
    parentId: Optional[str] = None
    image: Optional[str] = ""
    banner: Optional[str] = ""
    promoImage: Optional[str] = ""  # right-side mega-menu image
    description: Optional[str] = ""
    sortOrder: int = 0
    isActive: bool = True


class BrandIn(BaseModel):
    name: str
    slug: Optional[str] = None
    logo: Optional[str] = ""
    description: Optional[str] = ""
    isActive: bool = True


class TierPrice(BaseModel):
    minQty: int
    maxQty: Optional[int] = None
    price: float


class ProductIn(BaseModel):
    name: str = Field(min_length=1)
    slug: Optional[str] = None
    sku: Optional[str] = None
    hsn: Optional[str] = ""
    brandId: Optional[str] = None
    categoryId: str = Field(min_length=1)
    vendorId: Optional[str] = None
    description: str = ""
    shortDescription: str = ""
    specifications: dict = {}
    images: List[str] = []
    price: float = Field(gt=0)
    mrp: float
    tierPricing: List[TierPrice] = []
    moq: int = 1
    maxQty: Optional[int] = None
    stock: int = 0
    lowStockThreshold: int = 5
    gst: float = 18
    weight: float = 0
    isActive: bool = True
    isFeatured: bool = False
    tags: List[str] = []


class CartItemIn(BaseModel):
    productId: str
    quantity: int


class AddressIn(BaseModel):
    fullName: str = Field(min_length=2)
    phone: str = Field(min_length=7, max_length=15)
    line1: str = Field(min_length=3)
    line2: Optional[str] = ""
    city: str = Field(min_length=2)
    state: str = Field(min_length=2)
    pincode: str = Field(min_length=4, max_length=10)
    country: str = "India"
    gstin: Optional[str] = ""
    company: Optional[str] = ""
    isDefault: bool = False


class CheckoutIn(BaseModel):
    address: AddressIn
    paymentMethod: str  # cod | razorpay
    couponCode: Optional[str] = None
    notes: Optional[str] = ""


class RFQIn(BaseModel):
    productId: Optional[str] = None
    productName: str
    quantity: int
    targetPrice: Optional[float] = None
    deliveryLocation: str
    requiredBy: Optional[str] = None
    notes: Optional[str] = ""


class QuotationIn(BaseModel):
    rfqId: str
    price: float
    moq: int
    validity: Optional[str] = None
    shippingCharge: float = 0
    notes: Optional[str] = ""


class ReviewIn(BaseModel):
    productId: str
    rating: int = Field(ge=1, le=5)
    title: Optional[str] = ""
    body: Optional[str] = ""


class CouponIn(BaseModel):
    code: str
    type: str = "percentage"  # percentage | fixed
    value: float
    minOrder: float = 0
    maxDiscount: Optional[float] = None
    expiresAt: Optional[str] = None
    usageLimit: Optional[int] = None
    isActive: bool = True


# ============ AUTH ============
@api.post("/auth/register")
async def register(data: RegisterIn):
    if await db.users.find_one({"email": data.email.lower()}):
        raise HTTPException(400, "Email already registered")
    role = data.role if data.role in ("customer", "vendor") else "customer"
    uid = new_id()
    user = {
        "id": uid, "email": data.email.lower(), "password": hash_pw(data.password),
        "name": data.name, "phone": data.phone, "company": data.company, "gstin": data.gstin,
        "role": role, "isActive": True, "isApproved": role != "vendor",
        "createdAt": now(),
    }
    await db.users.insert_one(user)
    if role == "vendor":
        await db.vendors.insert_one({
            "id": new_id(), "userId": uid, "companyName": data.company or data.name,
            "email": data.email.lower(), "phone": data.phone, "gstin": data.gstin,
            "status": "pending", "commissionPct": 10, "walletBalance": 0,
            "createdAt": now(),
        })
    # welcome email (async, non-blocking)
    try:
        await send_email(data.email, "Welcome to TradeHub",
                         f"<h2>Welcome {data.name}!</h2><p>Your {role} account is ready.</p>")
    except Exception:
        pass
    token = make_access_token(uid, role)
    return {"token": token, "refresh": make_refresh_token(uid),
            "user": {"id": uid, "email": user["email"], "name": user["name"], "role": role}}


@api.post("/auth/login")
async def login(data: LoginIn):
    user = await db.users.find_one({"email": data.email.lower()})
    if not user or not verify_pw(data.password, user["password"]):
        raise HTTPException(401, "Invalid email or password")
    if not user.get("isActive", True):
        raise HTTPException(403, "Account disabled")
    token = make_access_token(user["id"], user["role"])
    return {"token": token, "refresh": make_refresh_token(user["id"]),
            "user": {"id": user["id"], "email": user["email"], "name": user["name"], "role": user["role"]}}


@api.get("/auth/me")
async def me(cur: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": cur["id"]}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(404, "User not found")
    return user


# ============ SETTINGS (branding / theme / features) ============
@api.get("/settings")
async def get_settings():
    s = await db.settings.find_one({"id": "global"}, {"_id": 0})
    if not s:
        s = {"id": "global", **DEFAULT_SETTINGS}
        await db.settings.insert_one({**s})
    return s


@api.put("/settings")
async def update_settings(data: dict, user: dict = Depends(require_permission("settings.manage"))):
    await db.settings.update_one({"id": "global"}, {"$set": data}, upsert=True)
    return await get_settings()


# ============ CATEGORIES ============
def slugify(s: str) -> str:
    return "".join(c.lower() if c.isalnum() else "-" for c in s).strip("-")


@api.get("/categories")
async def list_categories(tree: bool = False):
    cats = clean_list(await db.categories.find({"isActive": True}).sort("sortOrder", 1).to_list(500))
    if not tree:
        return cats
    by_id = {c["id"]: {**c, "children": []} for c in cats}
    roots = []
    for c in by_id.values():
        pid = c.get("parentId")
        if pid and pid in by_id:
            by_id[pid]["children"].append(c)
        else:
            roots.append(c)
    return roots


@api.post("/categories")
async def create_category(data: CategoryIn, user: dict = Depends(require_permission("settings.manage"))):
    d = data.model_dump()
    d["id"] = new_id()
    d["slug"] = d.get("slug") or slugify(d["name"])
    d["createdAt"] = now()
    await db.categories.insert_one(d)
    return clean(d)


@api.put("/categories/{cid}")
async def update_category(cid: str, data: dict, user: dict = Depends(require_permission("settings.manage"))):
    await db.categories.update_one({"id": cid}, {"$set": data})
    return clean(await db.categories.find_one({"id": cid}))


@api.delete("/categories/{cid}")
async def delete_category(cid: str, user: dict = Depends(require_permission("settings.manage"))):
    await db.categories.delete_one({"id": cid})
    return {"ok": True}


# ============ BRANDS ============
@api.get("/brands")
async def list_brands():
    return clean_list(await db.brands.find({"isActive": True}).to_list(500))


@api.post("/brands")
async def create_brand(data: BrandIn, user: dict = Depends(require_permission("settings.manage"))):
    d = data.model_dump()
    d["id"] = new_id()
    d["slug"] = d.get("slug") or slugify(d["name"])
    d["createdAt"] = now()
    await db.brands.insert_one(d)
    return clean(d)


@api.put("/brands/{bid}")
async def update_brand(bid: str, data: dict, user: dict = Depends(require_permission("settings.manage"))):
    await db.brands.update_one({"id": bid}, {"$set": data})
    return clean(await db.brands.find_one({"id": bid}))


@api.delete("/brands/{bid}")
async def delete_brand(bid: str, user: dict = Depends(require_permission("settings.manage"))):
    await db.brands.delete_one({"id": bid})
    return {"ok": True}


# ============ PRODUCTS ============
async def _enrich_products(prods: list) -> list:
    cat_ids = {p["categoryId"] for p in prods if p.get("categoryId")}
    brand_ids = {p.get("brandId") for p in prods if p.get("brandId")}
    vendor_ids = {p.get("vendorId") for p in prods if p.get("vendorId")}
    cats = {c["id"]: c for c in await db.categories.find({"id": {"$in": list(cat_ids)}}, {"_id": 0}).to_list(200)}
    brands = {b["id"]: b for b in await db.brands.find({"id": {"$in": list(brand_ids)}}, {"_id": 0}).to_list(200)}
    vendors = {v["id"]: v for v in await db.vendors.find({"id": {"$in": list(vendor_ids)}}, {"_id": 0}).to_list(200)}
    for p in prods:
        p["category"] = cats.get(p.get("categoryId"))
        p["brand"] = brands.get(p.get("brandId"))
        v = vendors.get(p.get("vendorId"))
        p["vendor"] = {"id": v["id"], "companyName": v["companyName"]} if v else None
    return prods


@api.get("/products")
async def list_products(
    q: Optional[str] = None,
    category: Optional[str] = None,
    brand: Optional[str] = None,
    vendor: Optional[str] = None,
    minPrice: Optional[float] = None,
    maxPrice: Optional[float] = None,
    featured: Optional[bool] = None,
    sort: str = "newest",
    page: int = 1,
    limit: int = 20,
):
    query = {"isActive": True}
    if q:
        query["$or"] = [{"name": {"$regex": q, "$options": "i"}}, {"sku": {"$regex": q, "$options": "i"}}]
    if category:
        query["categoryId"] = category
    if brand:
        query["brandId"] = brand
    if vendor:
        query["vendorId"] = vendor
    if featured is not None:
        query["isFeatured"] = featured
    if minPrice is not None or maxPrice is not None:
        pr = {}
        if minPrice is not None: pr["$gte"] = minPrice
        if maxPrice is not None: pr["$lte"] = maxPrice
        query["price"] = pr
    sort_map = {"newest": ("createdAt", -1), "price_low": ("price", 1),
                "price_high": ("price", -1), "popular": ("soldCount", -1)}
    field, direction = sort_map.get(sort, ("createdAt", -1))
    total = await db.products.count_documents(query)
    skip = (page - 1) * limit
    prods = clean_list(await db.products.find(query).sort(field, direction).skip(skip).limit(limit).to_list(limit))
    await _enrich_products(prods)
    return {"items": prods, "total": total, "page": page, "limit": limit}


@api.get("/products/{pid}")
async def get_product(pid: str):
    p = await db.products.find_one({"$or": [{"id": pid}, {"slug": pid}]}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Product not found")
    await _enrich_products([p])
    p["reviews"] = clean_list(await db.reviews.find({"productId": p["id"], "isApproved": True}).to_list(50))
    return p


@api.post("/products")
async def create_product(data: ProductIn, user: dict = Depends(get_current_user)):
    if user["role"] not in ("super_admin", "admin", "vendor"):
        raise HTTPException(403, "Not allowed")
    d = data.model_dump()
    d["id"] = new_id()
    d["slug"] = d.get("slug") or slugify(d["name"])
    d["sku"] = d.get("sku") or f"SKU-{d['id'][:8].upper()}"
    d["createdAt"] = now()
    d["soldCount"] = 0
    if user["role"] == "vendor":
        vendor = await db.vendors.find_one({"userId": user["id"]})
        if not vendor:
            raise HTTPException(400, "Vendor profile missing")
        d["vendorId"] = vendor["id"]
    await db.products.insert_one(d)
    return clean(d)


@api.put("/products/{pid}")
async def update_product(pid: str, data: dict, user: dict = Depends(get_current_user)):
    p = await db.products.find_one({"id": pid})
    if not p:
        raise HTTPException(404, "Not found")
    if user["role"] == "vendor":
        vendor = await db.vendors.find_one({"userId": user["id"]})
        if not vendor or p.get("vendorId") != vendor["id"]:
            raise HTTPException(403, "Not your product")
    elif user["role"] not in ("super_admin", "admin"):
        raise HTTPException(403, "Not allowed")
    await db.products.update_one({"id": pid}, {"$set": data})
    return clean(await db.products.find_one({"id": pid}))


@api.delete("/products/{pid}")
async def delete_product(pid: str, user: dict = Depends(get_current_user)):
    if user["role"] not in ("super_admin", "admin", "vendor"):
        raise HTTPException(403, "Not allowed")
    p = await db.products.find_one({"id": pid})
    if not p:
        raise HTTPException(404, "Not found")
    if user["role"] == "vendor":
        vendor = await db.vendors.find_one({"userId": user["id"]})
        if not vendor or p.get("vendorId") != vendor["id"]:
            raise HTTPException(403, "Not your product")
    await db.products.delete_one({"id": pid})
    return {"ok": True}


# ============ CART ============
async def _get_cart_doc(user_id: str):
    c = await db.carts.find_one({"userId": user_id})
    if not c:
        c = {"id": new_id(), "userId": user_id, "items": [], "createdAt": now()}
        await db.carts.insert_one(c)
    return c


async def _cart_summary(user_id: str):
    cart = await _get_cart_doc(user_id)
    items = []
    subtotal = 0
    tax_total = 0
    for it in cart.get("items", []):
        p = await db.products.find_one({"id": it["productId"]}, {"_id": 0})
        if not p:
            continue
        qty = it["quantity"]
        unit = price_for_qty(p.get("tierPricing", []), qty, p["price"])
        line = unit * qty
        line_tax = line * (p.get("gst", 0) / 100)
        items.append({
            "productId": p["id"], "name": p["name"], "image": (p.get("images") or [""])[0],
            "sku": p.get("sku"), "quantity": qty, "unitPrice": unit, "mrp": p.get("mrp"),
            "lineTotal": line, "lineTax": line_tax, "moq": p.get("moq", 1),
            "stock": p.get("stock", 0), "vendorId": p.get("vendorId"), "gst": p.get("gst", 0),
        })
        subtotal += line
        tax_total += line_tax
    shipping = 0 if subtotal >= 10000 else (99 if subtotal > 0 else 0)
    return {
        "items": items, "subtotal": subtotal, "tax": tax_total,
        "shipping": shipping, "total": subtotal + tax_total + shipping,
    }


@api.get("/cart")
async def get_cart(user: dict = Depends(get_current_user)):
    return await _cart_summary(user["id"])


@api.post("/cart/add")
async def cart_add(data: CartItemIn, user: dict = Depends(get_current_user)):
    p = await db.products.find_one({"id": data.productId})
    if not p:
        raise HTTPException(404, "Product not found")
    if data.quantity < p.get("moq", 1):
        raise HTTPException(400, f"Minimum order quantity is {p.get('moq', 1)}")
    cart = await _get_cart_doc(user["id"])
    items = cart.get("items", [])
    found = False
    for it in items:
        if it["productId"] == data.productId:
            it["quantity"] = data.quantity
            found = True
            break
    if not found:
        items.append({"productId": data.productId, "quantity": data.quantity})
    await db.carts.update_one({"userId": user["id"]}, {"$set": {"items": items}})
    return await _cart_summary(user["id"])


@api.delete("/cart/{pid}")
async def cart_remove(pid: str, user: dict = Depends(get_current_user)):
    cart = await _get_cart_doc(user["id"])
    items = [i for i in cart.get("items", []) if i["productId"] != pid]
    await db.carts.update_one({"userId": user["id"]}, {"$set": {"items": items}})
    return await _cart_summary(user["id"])


# ============ ORDERS ============
async def _apply_coupon(code: str, subtotal: float):
    if not code:
        return 0, None
    c = await db.coupons.find_one({"code": code.upper(), "isActive": True})
    if not c:
        return 0, None
    if subtotal < c.get("minOrder", 0):
        return 0, None
    if c["type"] == "percentage":
        disc = subtotal * c["value"] / 100
        if c.get("maxDiscount"):
            disc = min(disc, c["maxDiscount"])
    else:
        disc = c["value"]
    return disc, c


@api.post("/orders/checkout")
async def checkout(data: CheckoutIn, user: dict = Depends(get_current_user)):
    summary = await _cart_summary(user["id"])
    if not summary["items"]:
        raise HTTPException(400, "Cart is empty")

    # server-side price recompute (never trust client)
    discount, coupon = await _apply_coupon(data.couponCode or "", summary["subtotal"])
    total = max(0, summary["subtotal"] + summary["tax"] + summary["shipping"] - discount)

    settings = await db.settings.find_one({"id": "global"}) or DEFAULT_SETTINGS
    prefix = settings.get("commerce", {}).get("orderPrefix", "ORD")
    order_no = f"{prefix}-{int(datetime.now().timestamp())}-{uuid.uuid4().hex[:6].upper()}"
    order_id = new_id()

    order = {
        "id": order_id, "orderNo": order_no, "userId": user["id"],
        "items": summary["items"], "address": data.address.model_dump(),
        "subtotal": summary["subtotal"], "tax": summary["tax"],
        "shipping": summary["shipping"], "discount": discount,
        "total": total, "couponCode": data.couponCode,
        "paymentMethod": data.paymentMethod, "notes": data.notes,
        "status": "pending_payment" if data.paymentMethod == "razorpay" else "confirmed",
        "paymentStatus": "pending",
        "timeline": [{"status": "created", "at": now()}],
        "createdAt": now(),
    }

    razor_order = None
    if data.paymentMethod == "razorpay":
        try:
            razor_order = await razor_create_order(
                int(total * 100), order_no,
                {"orderId": order_id, "userId": user["id"]},
            )
            order["razorpayOrderId"] = razor_order["id"]
        except Exception as e:
            log.warning(f"razorpay order create failed: {e}")
            raise HTTPException(502, f"Payment gateway error: {e}")
    else:
        order["paymentStatus"] = "cod_pending"
        order["timeline"].append({"status": "confirmed", "at": now()})

    await db.orders.insert_one(order)

    # For prepaid (razorpay): don't decrement stock or clear cart until payment is verified.
    if data.paymentMethod != "razorpay":
        for it in summary["items"]:
            await db.products.update_one({"id": it["productId"]}, {"$inc": {"stock": -it["quantity"], "soldCount": it["quantity"]}})
        await db.carts.update_one({"userId": user["id"]}, {"$set": {"items": []}})

    # email confirmation
    u = await db.users.find_one({"id": user["id"]})
    try:
        await send_email(u["email"], f"Order {order_no} received",
                         f"<h3>Thanks {u['name']}</h3><p>Order <b>{order_no}</b> for ₹{total:.2f} received.</p>")
    except Exception:
        pass

    return {
        "order": clean(order),
        "razorpay": {
            "orderId": razor_order["id"], "amount": int(total * 100),
            "currency": "INR", "keyId": RAZORPAY_KEY_ID,
        } if razor_order else None,
    }


class PaymentVerifyIn(BaseModel):
    orderId: str
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str


@api.post("/orders/verify-payment")
async def verify_payment(data: PaymentVerifyIn, user: dict = Depends(get_current_user)):
    ok = await razor_verify_signature(data.razorpay_order_id, data.razorpay_payment_id, data.razorpay_signature)
    if not ok:
        raise HTTPException(400, "Invalid payment signature")
    order = await db.orders.find_one({"id": data.orderId, "userId": user["id"]})
    if not order:
        raise HTTPException(404, "Order not found")
    if order.get("paymentStatus") != "paid":
        # Now decrement stock + clear cart (post-verification)
        for it in order.get("items", []):
            await db.products.update_one({"id": it["productId"]}, {"$inc": {"stock": -it["quantity"], "soldCount": it["quantity"]}})
        await db.carts.update_one({"userId": user["id"]}, {"$set": {"items": []}})
    await db.orders.update_one({"id": data.orderId, "userId": user["id"]}, {"$set": {
        "paymentStatus": "paid", "status": "confirmed",
        "razorpayPaymentId": data.razorpay_payment_id, "paidAt": now(),
    }, "$push": {"timeline": {"status": "paid", "at": now()}}})
    return {"ok": True}


@api.get("/orders")
async def my_orders(user: dict = Depends(get_current_user)):
    if user["role"] in ("super_admin", "admin"):
        docs = await db.orders.find({}, {"_id": 0}).sort("createdAt", -1).to_list(500)
    elif user["role"] == "vendor":
        vendor = await db.vendors.find_one({"userId": user["id"]})
        vid = vendor["id"] if vendor else None
        docs = await db.orders.find({"items.vendorId": vid}, {"_id": 0}).sort("createdAt", -1).to_list(500)
    else:
        docs = await db.orders.find({"userId": user["id"]}, {"_id": 0}).sort("createdAt", -1).to_list(500)
    return docs


@api.get("/orders/{oid}")
async def get_order(oid: str, user: dict = Depends(get_current_user)):
    o = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not o:
        raise HTTPException(404, "Not found")
    if user["role"] not in ("super_admin", "admin") and o["userId"] != user["id"]:
        if user["role"] == "vendor":
            vendor = await db.vendors.find_one({"userId": user["id"]})
            if not vendor or not any(i.get("vendorId") == vendor["id"] for i in o["items"]):
                raise HTTPException(403, "Forbidden")
        else:
            raise HTTPException(403, "Forbidden")
    return o


@api.put("/orders/{oid}/status")
async def update_order_status(oid: str, data: dict, user: dict = Depends(get_current_user)):
    if user["role"] not in ("super_admin", "admin", "vendor"):
        raise HTTPException(403, "Not allowed")
    status = data.get("status")
    if not status:
        raise HTTPException(400, "status required")
    order = await db.orders.find_one({"id": oid})
    if not order:
        raise HTTPException(404, "Order not found")
    prev = order.get("status")
    await db.orders.update_one({"id": oid}, {
        "$set": {"status": status},
        "$push": {"timeline": {"status": status, "at": now(), "by": user["id"]}},
    })
    # On first transition to delivered → credit vendor wallets with commission deducted
    if status == "delivered" and prev != "delivered":
        by_vendor = {}
        for it in order.get("items", []):
            vid = it.get("vendorId")
            if not vid:
                continue
            by_vendor.setdefault(vid, 0.0)
            by_vendor[vid] += float(it.get("lineTotal", 0))
        for vid, gross in by_vendor.items():
            v = await db.vendors.find_one({"id": vid})
            if not v:
                continue
            pct = float(v.get("commissionPct", 10))
            commission = round(gross * pct / 100, 2)
            net = round(gross - commission, 2)
            await db.vendors.update_one({"id": vid}, {"$inc": {"walletBalance": net}})
            await db.wallet_transactions.insert_one({
                "id": new_id(), "vendorId": vid, "orderId": oid, "orderNo": order.get("orderNo"),
                "type": "credit", "gross": gross, "commissionPct": pct, "commission": commission,
                "amount": net, "balance": (v.get("walletBalance", 0) + net),
                "status": "available", "note": f"Order {order.get('orderNo')} delivered",
                "createdAt": now(),
            })
    return clean(await db.orders.find_one({"id": oid}))


@api.post("/orders/{oid}/ship")
async def ship_order(oid: str, user: dict = Depends(get_current_user)):
    if user["role"] not in ("super_admin", "admin", "vendor"):
        raise HTTPException(403, "Not allowed")
    o = await db.orders.find_one({"id": oid})
    if not o:
        raise HTTPException(404, "Not found")
    addr = o["address"]
    payload = {
        "order_id": o["orderNo"],
        "order_date": o["createdAt"][:10],
        "pickup_location": "Primary",
        "billing_customer_name": addr["fullName"],
        "billing_address": addr["line1"],
        "billing_city": addr["city"],
        "billing_pincode": addr["pincode"],
        "billing_state": addr["state"],
        "billing_country": addr.get("country", "India"),
        "billing_email": (await db.users.find_one({"id": o["userId"]}) or {}).get("email", "noreply@example.com"),
        "billing_phone": addr["phone"],
        "shipping_is_billing": True,
        "order_items": [{"name": i["name"], "sku": i.get("sku", ""), "units": i["quantity"],
                          "selling_price": i["unitPrice"]} for i in o["items"]],
        "payment_method": "COD" if o["paymentMethod"] == "cod" else "Prepaid",
        "sub_total": o["subtotal"], "length": 10, "breadth": 10, "height": 10, "weight": 1,
    }
    result = await ship_create_order(payload)
    await db.orders.update_one({"id": oid}, {"$set": {
        "status": "shipped", "shiprocket": result,
        "shipmentId": result.get("shipment_id"), "awb": result.get("awb_code"),
    }, "$push": {"timeline": {"status": "shipped", "at": now()}}})
    return {"result": result}


# ============ VENDORS ============
@api.get("/vendors")
async def list_vendors(status: Optional[str] = None):
    q = {}
    if status:
        q["status"] = status
    return clean_list(await db.vendors.find(q).to_list(500))


@api.get("/vendors/me")
async def my_vendor(user: dict = Depends(require_roles("vendor"))):
    v = await db.vendors.find_one({"userId": user["id"]}, {"_id": 0})
    if not v:
        raise HTTPException(404, "Vendor profile not found")
    return v


@api.put("/vendors/{vid}/approve")
async def approve_vendor(vid: str, data: dict, user: dict = Depends(require_permission("vendors.approve"))):
    status = data.get("status", "approved")
    await db.vendors.update_one({"id": vid}, {"$set": {"status": status}})
    return {"ok": True}


# ============ RFQ / QUOTATIONS ============
@api.post("/rfq")
async def create_rfq(data: RFQIn, user: dict = Depends(get_current_user)):
    d = data.model_dump()
    d["id"] = new_id()
    d["userId"] = user["id"]
    d["status"] = "open"
    d["createdAt"] = now()
    await db.rfqs.insert_one(d)
    return clean(d)


@api.get("/rfq")
async def list_rfq(user: dict = Depends(get_current_user)):
    if user["role"] in ("super_admin", "admin"):
        docs = await db.rfqs.find({}, {"_id": 0}).sort("createdAt", -1).to_list(200)
    elif user["role"] == "vendor":
        docs = await db.rfqs.find({"status": "open"}, {"_id": 0}).sort("createdAt", -1).to_list(200)
    else:
        docs = await db.rfqs.find({"userId": user["id"]}, {"_id": 0}).sort("createdAt", -1).to_list(200)
    for r in docs:
        r["quotations"] = clean_list(await db.quotations.find({"rfqId": r["id"]}).to_list(50))
    return docs


@api.post("/quotations")
async def create_quotation(data: QuotationIn, user: dict = Depends(require_roles("vendor"))):
    v = await db.vendors.find_one({"userId": user["id"]})
    d = data.model_dump()
    d["id"] = new_id()
    d["vendorId"] = v["id"] if v else None
    d["vendorName"] = v["companyName"] if v else "Vendor"
    d["status"] = "pending"
    d["createdAt"] = now()
    await db.quotations.insert_one(d)
    return clean(d)


@api.post("/quotations/{qid}/accept")
async def accept_quotation(qid: str, user: dict = Depends(get_current_user)):
    await db.quotations.update_one({"id": qid}, {"$set": {"status": "accepted"}})
    return {"ok": True}


# ============ COUPONS ============
@api.get("/coupons")
async def list_coupons(user: dict = Depends(get_current_user)):
    return clean_list(await db.coupons.find({}, {"_id": 0}).to_list(200))


@api.post("/coupons")
async def create_coupon(data: CouponIn, user: dict = Depends(require_permission("coupons.manage"))):
    d = data.model_dump()
    d["code"] = d["code"].upper()
    d["id"] = new_id()
    d["createdAt"] = now()
    await db.coupons.insert_one(d)
    return clean(d)


@api.post("/coupons/apply")
async def apply_coupon(data: dict, user: dict = Depends(get_current_user)):
    summary = await _cart_summary(user["id"])
    disc, c = await _apply_coupon(data.get("code", ""), summary["subtotal"])
    if not c:
        raise HTTPException(400, "Invalid or ineligible coupon")
    return {"discount": disc, "coupon": clean(c)}


# ============ REVIEWS ============
@api.post("/reviews")
async def create_review(data: ReviewIn, user: dict = Depends(get_current_user)):
    d = data.model_dump()
    d["id"] = new_id()
    d["userId"] = user["id"]
    u = await db.users.find_one({"id": user["id"]})
    d["userName"] = u["name"] if u else "Customer"
    d["isApproved"] = True
    d["createdAt"] = now()
    await db.reviews.insert_one(d)
    return clean(d)


# ============ MEDIA UPLOAD ============
@api.post("/media/upload")
async def upload_file(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
        raise HTTPException(400, "Unsupported file type")
    fname = f"{new_id()}{ext}"
    path = UPLOAD_DIR / fname
    content = await file.read()
    if len(content) > 8 * 1024 * 1024:
        raise HTTPException(400, "File too large (max 8MB)")
    path.write_bytes(content)
    return {"url": f"/api/media/{fname}", "filename": fname}


@api.get("/media/{filename}")
async def serve_media(filename: str):
    path = UPLOAD_DIR / filename
    if not path.exists():
        raise HTTPException(404, "Not found")
    return FileResponse(str(path))


# ============ ADMIN DASHBOARD ============
@api.get("/admin/stats")
async def admin_stats(user: dict = Depends(require_permission("dashboard.view"))):
    total_orders = await db.orders.count_documents({})
    paid_orders = await db.orders.count_documents({"paymentStatus": "paid"})
    cursor = db.orders.aggregate([
        {"$group": {"_id": None, "revenue": {"$sum": "$total"}, "count": {"$sum": 1}}}
    ])
    agg = await cursor.to_list(1)
    revenue = agg[0]["revenue"] if agg else 0
    return {
        "totalOrders": total_orders,
        "paidOrders": paid_orders,
        "revenue": revenue,
        "totalCustomers": await db.users.count_documents({"role": "customer"}),
        "totalVendors": await db.vendors.count_documents({}),
        "pendingVendors": await db.vendors.count_documents({"status": "pending"}),
        "totalProducts": await db.products.count_documents({}),
        "activeProducts": await db.products.count_documents({"isActive": True}),
        "lowStock": await db.products.count_documents({"$expr": {"$lte": ["$stock", "$lowStockThreshold"]}}),
    }


@api.get("/admin/users")
async def admin_users(user: dict = Depends(require_permission("customers.view"))):
    return clean_list(await db.users.find({}, {"_id": 0, "password": 0}).to_list(500))


# ============ WEBHOOKS ============
@api.post("/webhooks/razorpay")
async def razorpay_webhook(request: Request):
    raw = await request.body()
    sig = request.headers.get("x-razorpay-signature", "")
    if not razor_verify_webhook(raw, sig):
        raise HTTPException(400, "Invalid signature")
    import json as _json
    event = _json.loads(raw)
    kind = event.get("event", "")
    payment = event.get("payload", {}).get("payment", {}).get("entity", {})
    if kind == "payment.captured" and payment.get("order_id"):
        await db.orders.update_one({"razorpayOrderId": payment["order_id"]}, {"$set": {"paymentStatus": "paid", "status": "confirmed"}})
    return {"ok": True}


# ============ BULK IMPORT / WALLET / PAYOUTS / FLASH SALES ============
import csv as _csv
from io import BytesIO, StringIO
from openpyxl import load_workbook

REQUIRED_COLS = ["name", "categoryId", "price", "mrp", "stock", "moq", "gst"]
OPTIONAL_COLS = ["sku", "brandId", "description", "hsn", "images", "isActive", "isFeatured"]


def _parse_rows(filename: str, content: bytes) -> list:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = _csv.DictReader(StringIO(text))
        return [dict(r) for r in reader]
    if name.endswith((".xlsx", ".xls")):
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter, [])]
        out = []
        for row in rows_iter:
            if all(v is None or v == "" for v in row):
                continue
            out.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
        return out
    raise HTTPException(400, "Unsupported file. Use .csv or .xlsx")


async def _validate_import(rows: list) -> dict:
    cats = {c["id"] for c in await db.categories.find({}, {"id": 1}).to_list(1000)}
    brands = {b["id"] for b in await db.brands.find({}, {"id": 1}).to_list(1000)}
    valid, errors = [], []
    seen_skus = set()
    for idx, r in enumerate(rows, start=2):  # header is row 1
        row_err = []
        for col in REQUIRED_COLS:
            if r.get(col) in (None, ""):
                row_err.append(f"{col} required")
        try:
            price = float(r.get("price") or 0)
            mrp = float(r.get("mrp") or price)
            stock = int(float(r.get("stock") or 0))
            moq = int(float(r.get("moq") or 1))
            gst = float(r.get("gst") or 18)
            if price <= 0:
                row_err.append("price must be > 0")
        except Exception:
            row_err.append("price/mrp/stock/moq/gst must be numeric")
            price = mrp = 0; stock = moq = 0; gst = 18
        cat = str(r.get("categoryId") or "").strip()
        if cat and cat not in cats:
            row_err.append(f"unknown categoryId '{cat}'")
        brand = str(r.get("brandId") or "").strip()
        if brand and brand not in brands:
            row_err.append(f"unknown brandId '{brand}'")
        sku = str(r.get("sku") or "").strip()
        if sku:
            if sku in seen_skus:
                row_err.append(f"duplicate SKU in file: {sku}")
            else:
                seen_skus.add(sku)
        item = {
            "row": idx,
            "data": {
                "name": str(r.get("name") or "").strip(),
                "sku": sku or None,
                "categoryId": cat,
                "brandId": brand or None,
                "description": str(r.get("description") or ""),
                "hsn": str(r.get("hsn") or ""),
                "price": price, "mrp": mrp if mrp else price,
                "stock": stock, "moq": moq, "gst": gst,
                "images": [i.strip() for i in str(r.get("images") or "").split("|") if i.strip()],
                "isActive": str(r.get("isActive") or "true").lower() in ("1", "true", "yes"),
                "isFeatured": str(r.get("isFeatured") or "false").lower() in ("1", "true", "yes"),
            },
            "errors": row_err,
        }
        (errors if row_err else valid).append(item)
    return {"valid": valid, "errors": errors, "total": len(rows)}


@api.get("/products/import/template")
async def import_template():
    header = REQUIRED_COLS + OPTIONAL_COLS
    sample = ["Sample Widget", "c-electronics", "199", "249", "100", "10", "18",
              "SKU-SAMPLE-01", "b-1", "Sample description", "8471", "https://example.com/img.jpg|https://example.com/img2.jpg", "true", "false"]
    body = ",".join(header) + "\n" + ",".join(sample) + "\n"
    from fastapi.responses import Response
    return Response(content=body, media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=product-import-template.csv"})


@api.post("/products/import/preview")
async def import_preview(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    if user["role"] not in ("super_admin", "admin", "vendor"):
        raise HTTPException(403, "Not allowed")
    content = await file.read()
    rows = _parse_rows(file.filename or "", content)
    return await _validate_import(rows)


@api.post("/products/import/commit")
async def import_commit(payload: dict, user: dict = Depends(get_current_user)):
    if user["role"] not in ("super_admin", "admin", "vendor"):
        raise HTTPException(403, "Not allowed")
    items = payload.get("items") or []
    vendor_id = None
    if user["role"] == "vendor":
        v = await db.vendors.find_one({"userId": user["id"]})
        vendor_id = v["id"] if v else None
    inserted = 0
    failed = []
    for it in items:
        d = it["data"] if "data" in it else it
        try:
            d["id"] = new_id()
            d["slug"] = slugify(d["name"]) + "-" + d["id"][:6]
            d["sku"] = d.get("sku") or f"SKU-{d['id'][:8].upper()}"
            d["vendorId"] = vendor_id or d.get("vendorId")
            d["createdAt"] = now()
            d["soldCount"] = 0
            d["tierPricing"] = d.get("tierPricing", [])
            d["specifications"] = d.get("specifications", {})
            d["lowStockThreshold"] = d.get("lowStockThreshold", 5)
            await db.products.insert_one(d)
            inserted += 1
        except Exception as e:
            failed.append({"row": it.get("row"), "error": str(e)})
    return {"inserted": inserted, "failed": failed}


# ---- Wallet ----
@api.get("/wallet/vendor")
async def vendor_wallet(user: dict = Depends(require_roles("vendor"))):
    v = await db.vendors.find_one({"userId": user["id"]}, {"_id": 0})
    if not v:
        raise HTTPException(404, "Vendor not found")
    txns = clean_list(await db.wallet_transactions.find({"vendorId": v["id"]}).sort("createdAt", -1).to_list(500))
    payouts = clean_list(await db.payouts.find({"vendorId": v["id"]}).sort("createdAt", -1).to_list(200))
    return {"vendor": v, "transactions": txns, "payouts": payouts}


class PayoutRequestIn(BaseModel):
    amount: float
    method: str = "bank_transfer"
    notes: Optional[str] = ""


@api.post("/payouts/request")
async def request_payout(data: PayoutRequestIn, user: dict = Depends(require_roles("vendor"))):
    v = await db.vendors.find_one({"userId": user["id"]})
    if not v:
        raise HTTPException(404, "Vendor not found")
    bal = float(v.get("walletBalance", 0))
    if data.amount <= 0 or data.amount > bal:
        raise HTTPException(400, f"Invalid amount. Available balance: ₹{bal}")
    p = {
        "id": new_id(), "vendorId": v["id"], "vendorName": v["companyName"],
        "amount": data.amount, "method": data.method, "notes": data.notes,
        "status": "pending", "createdAt": now(),
    }
    await db.payouts.insert_one(p)
    return clean(p)


@api.get("/payouts")
async def list_payouts(user: dict = Depends(get_current_user)):
    if user["role"] in ("super_admin", "admin"):
        docs = await db.payouts.find({}, {"_id": 0}).sort("createdAt", -1).to_list(500)
    elif user["role"] == "vendor":
        v = await db.vendors.find_one({"userId": user["id"]})
        docs = await db.payouts.find({"vendorId": v["id"] if v else "_"}, {"_id": 0}).sort("createdAt", -1).to_list(500)
    else:
        raise HTTPException(403, "Not allowed")
    return docs


@api.put("/payouts/{pid}/status")
async def update_payout(pid: str, data: dict, user: dict = Depends(require_permission("payments.manage"))):
    status = data.get("status")
    if status not in ("approved", "paid", "rejected"):
        raise HTTPException(400, "invalid status")
    p = await db.payouts.find_one({"id": pid})
    if not p:
        raise HTTPException(404, "Not found")
    updates = {"status": status, "updatedAt": now()}
    if status == "paid":
        updates["paidAt"] = now()
        # deduct from wallet + record txn
        await db.vendors.update_one({"id": p["vendorId"]}, {"$inc": {"walletBalance": -float(p["amount"])}})
        await db.wallet_transactions.insert_one({
            "id": new_id(), "vendorId": p["vendorId"], "type": "debit",
            "amount": -float(p["amount"]), "status": "paid",
            "note": f"Payout {pid} paid via {p.get('method')}", "createdAt": now(),
        })
    await db.payouts.update_one({"id": pid}, {"$set": updates})
    return clean(await db.payouts.find_one({"id": pid}))


# ---- Flash Sales ----
class FlashSaleIn(BaseModel):
    name: str
    discountPct: float = Field(gt=0, le=90)
    productIds: List[str] = []
    categoryId: Optional[str] = None
    startsAt: str
    endsAt: str
    banner: Optional[str] = ""
    isActive: bool = True


@api.get("/flash-sales")
async def list_flash_sales(activeOnly: bool = False):
    q = {}
    if activeOnly:
        n = now()
        q = {"isActive": True, "startsAt": {"$lte": n}, "endsAt": {"$gte": n}}
    return clean_list(await db.flash_sales.find(q).sort("startsAt", -1).to_list(200))


@api.post("/flash-sales")
async def create_flash_sale(data: FlashSaleIn, user: dict = Depends(require_permission("coupons.manage"))):
    d = data.model_dump()
    d["id"] = new_id()
    d["createdAt"] = now()
    await db.flash_sales.insert_one(d)
    return clean(d)


@api.put("/flash-sales/{fid}")
async def update_flash_sale(fid: str, data: dict, user: dict = Depends(require_permission("coupons.manage"))):
    await db.flash_sales.update_one({"id": fid}, {"$set": data})
    return clean(await db.flash_sales.find_one({"id": fid}))


@api.delete("/flash-sales/{fid}")
async def delete_flash_sale(fid: str, user: dict = Depends(require_permission("coupons.manage"))):
    await db.flash_sales.delete_one({"id": fid})
    return {"ok": True}


@api.get("/orders/{oid}/invoice")
async def download_invoice(oid: str, user: dict = Depends(get_current_user)):
    from fastapi.responses import StreamingResponse
    from invoice_pdf import build_invoice_pdf
    order = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    if user["role"] not in ("super_admin", "admin"):
        if order.get("userId") != user["id"]:
            if user["role"] == "vendor":
                vendor = await db.vendors.find_one({"userId": user["id"]})
                if not vendor or not any(i.get("vendorId") == vendor["id"] for i in order.get("items", [])):
                    raise HTTPException(403, "Forbidden")
            else:
                raise HTTPException(403, "Forbidden")
    settings = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {}
    buyer = await db.users.find_one({"id": order["userId"]}, {"_id": 0, "password": 0}) or {}
    pdf = await build_invoice_pdf(order, settings, buyer)
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="Invoice-{order.get("orderNo")}.pdf"'},
    )


# ============ HEALTH ============
@api.get("/")
async def root():
    return {"message": "B2B Marketplace API", "version": "1.0"}


app.include_router(api)


@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.orders.create_index("orderNo", unique=True)
    await db.products.create_index("slug", unique=True)
    log.info("Server started")


@app.on_event("shutdown")
async def shutdown():
    client.close()
